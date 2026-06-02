import openpyxl, json
wb = openpyxl.load_workbook("QA Transports URLs.xlsx", data_only=True)
ws = wb["Transports"]
print(f"Transports: {ws.max_row} filas x {ws.max_column} columnas")
# Headers (row 4)
headers = [str(c.value) if c.value else f"col{i}" for i, c in enumerate(ws[4])]
print("HEADERS:", headers)
# Sample 5 rows with FULL values
for row in ws.iter_rows(min_row=5, max_row=9, values_only=True):
    d = {h: str(v) if v else "" for h, v in zip(headers, row)}
    print(json.dumps(d, indent=2, ensure_ascii=False))
# Count unique projects
projects = set()
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    if row[0]: projects.add(row[0])
print(f"\nProyectos únicos: {projects}")
# Count rows with CAB
cab_vals = set()
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    if row[-1] is not None: cab_vals.add(row[-1])
print(f"CAB values: {cab_vals}")
